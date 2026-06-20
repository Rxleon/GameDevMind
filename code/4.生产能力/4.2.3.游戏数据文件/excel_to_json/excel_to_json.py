#!/usr/bin/env python3
"""
excel_to_json.py — Excel 策划表 → JSON 流水线最小示例

对应文档：mds/4.生产能力/4.2.3.游戏数据文件.md
依赖：openpyxl（pip install openpyxl）
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any, Dict, List

try:
    from openpyxl import load_workbook
except ImportError:
    print("请先安装：pip install openpyxl")
    sys.exit(1)


def parse_type(value: Any, col_type: str) -> Any:
    if value is None or value == "":
        return None
    if col_type == "int":
        return int(value)
    if col_type == "float":
        return float(value)
    if col_type == "bool":
        return str(value).lower() in ("1", "true", "yes", "y")
    return str(value).strip()


def excel_to_json(xlsx_path: Path, sheet_name: str | None = None) -> List[Dict[str, Any]]:
    """
    约定 Excel 格式（与多数手游策划表一致）：
      第 1 行：字段名（英文 key）
      第 2 行：类型（int/float/string/bool）
      第 3 行：注释（导出时忽略）
      第 4 行起：数据
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        raise ValueError("至少需要 4 行：字段名、类型、注释、数据")

    keys = [str(c).strip() for c in rows[0] if c is not None]
    types = [str(c).strip().lower() if c else "string" for c in rows[1][: len(keys)]]
    records: List[Dict[str, Any]] = []

    for row in rows[3:]:
        if not row or row[0] is None:
            continue
        item: Dict[str, Any] = {}
        for i, key in enumerate(keys):
            raw = row[i] if i < len(row) else None
            item[key] = parse_type(raw, types[i])
        records.append(item)

    wb.close()
    return records


def export(xlsx: Path, out_json: Path, id_field: str = "id") -> None:
    data = excel_to_json(xlsx)
    # 常见模式：按 id 索引的字典，客户端 O(1) 查找
    indexed = {str(row[id_field]): row for row in data if row.get(id_field) is not None}
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(indexed, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"导出 {len(indexed)} 条 → {out_json}")


if __name__ == "__main__":
    base = Path(__file__).resolve().parent
    xlsx = base / "sample" / "hero_config.xlsx"
    out = base / "output" / "hero_config.json"
    export(xlsx, out)
